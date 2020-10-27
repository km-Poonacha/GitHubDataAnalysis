# -*- coding: utf-8 -*-
"""
Created on Wed Jan  2 10:30:42 2019

@author: kmpoo

Aggregate commit creativity at the level of months. Remove projects for which no commits were found. 
"""

import pandas as pd
import numpy as np
import ast



REPOCOMMIT_LIST =[                   
                    r"C:\Users\pmedappa\Dropbox\Data\092019 CommitInfo\ClassifiedRepoCommit\ClassifiedRepoCommit1_287_1.xlsx",
                    "C:/Users/pmedappa/Dropbox/Data/092019 CommitInfo/ClassifiedRepoCommit/ClassifiedRepoCommit288_500_1.xlsx",
                    "C:/Users/pmedappa/Dropbox/Data/092019 CommitInfo/ClassifiedRepoCommit/ClassifiedRepoCommit501_1000_1.xlsx",
                    "C:/Users/pmedappa/Dropbox/Data/092019 CommitInfo/ClassifiedRepoCommit/ClassifiedRepoCommit1001_1500_1.xlsx",
                    "C:/Users/pmedappa/Dropbox/Data/092019 CommitInfo/ClassifiedRepoCommit/ClassifiedRepoCommit1501_2000_1.xlsx",
                    "C:/Users/pmedappa/Dropbox/Data/092019 CommitInfo/ClassifiedRepoCommit/ClassifiedRepoCommit2002_2500_1.xlsx",
                    "C:/Users/pmedappa/Dropbox/Data/092019 CommitInfo/ClassifiedRepoCommit/ClassifiedRepoCommit2501_3250_1.xlsx",
                    "C:/Users/pmedappa/Dropbox/Data/092019 CommitInfo/ClassifiedRepoCommit/ClassifiedRepoCommit3251_4000_1.xlsx",
                    "C:/Users/pmedappa/Dropbox/Data/092019 CommitInfo/ClassifiedRepoCommit/ClassifiedRepoCommit4001_5000_1.xlsx",
                    "C:/Users/pmedappa/Dropbox/Data/092019 CommitInfo/ClassifiedRepoCommit/ClassifiedRepoCommit5001_5202_1.xlsx",
                    "C:/Users/pmedappa/Dropbox/Data/092019 CommitInfo/ClassifiedRepoCommit/ClassifiedRepoCommit5203_6000_1.xlsx",
                    "C:/Users/pmedappa/Dropbox/Data/092019 CommitInfo/ClassifiedRepoCommit/ClassifiedRepoCommit6001_6570_1.xlsx"
                  ] 
MC_XLSX = "C:/Users/pmedappa/Dropbox/Data/092019 CommitInfo/ClassifiedRepoCommit/28072020_MC_RepoCommit.xlsx"
COMMITERS_XLSX = "C:/Users/pmedappa/Dropbox/Data/092019 CommitInfo/Contributors_monthwise/Final_COL_MC_RepoCommit.xlsx"
TEMP= "C:/Users/pmedappa/Dropbox/Data/092019 CommitInfo/ClassifiedRepoCommit/temp.xlsx"
def consolidate_prob(x, a1, a2,a3):
    "Aggregae the probabilities calculated into a single construct"
    text_file = ['txt','md','doc','docx','png','gif','jpg',"csv","xlsx"]
    if pd.isna(x['PINDEX']) and pd.notna(x['OPEN_ISSUES']):
        try:
            # Check for EOL issues
            files = ast.literal_eval(x['OPEN_ISSUES'])
        except:
            return np.NaN

        for i in files:
            l = i.split('.')
            if len(l) > 1:
                if l[1].lower() not in text_file:
                    return (x[a1]+x[a2]*2+x[a3]*3)/3

    return np.NaN

def consolidate_commits(df):
    """ Consolidate commits monthly"""
    if int(df.shape[0]) <1: return pd.DataFrame()
    df.dropna(subset=['con_novelty'],inplace = True)
    df=df.replace('Not Found', 0)
    df=df.fillna(0)

    df['documentation'] = pd.Series([1 if (('readme' in str(x).lower()) or ('document' in str(x).lower())) else 0 for x in df['MAIN_LANGUAGE']])
    df['peer'] = pd.Series([1 if 'merge' in str(x).lower() else 0 for x in df['MAIN_LANGUAGE']])
    df['bug'] = pd.Series([1 if (('bug' in str(x).lower()) or ('fix' in str(x).lower())) else 0 for x in df['MAIN_LANGUAGE']])
    df['feature'] = pd.Series([1 if (('feature' in str(x).lower()) or ('patch' in str(x).lower()) or ('update' in str(x).lower()) or ('function' in str(x).lower())) else 0 for x in df['MAIN_LANGUAGE']])
    df['test'] = pd.Series([1 if (('test' in str(x).lower()) or ('debug' in str(x).lower()) or ('check' in str(x).lower())) else 0 for x in df['MAIN_LANGUAGE']])

    df2 = df.groupby('c_month')['con_novelty', 'con_usefulness','Noveltys2p1','Noveltys2p2','Noveltys2p3','Usefulnesss2p1','Usefulnesss2p2','Usefulnesss2p3','NO_LANGUAGES','SCRIPT_SIZE','STARS','SUBSCRIPTIONS'].mean()
    df2= pd.concat([df2, df.groupby('c_month')['con_novelty'].count(),df.groupby('c_month')['SIZE'].nunique(),df.groupby('c_month')['documentation','peer','bug','feature','test'].sum()], axis = 1)
    # Change index from c_months 
    
    df2= df2.reset_index()
    if df2.shape[0] != 0:
        df2.columns = ['nc_month','m_novelty','m_usefulness','m_Noveltys2p1','m_Noveltys2p2','m_Noveltys2p3','m_Usefulnesss2p1','m_Usefulnesss2p2','m_Usefulnesss2p3','mavg_added','mavg_deleted','mavg_parents','mavg_files','c_count','c_contributors','documentation','peer','bug','feature','test']
    else:
        df2 = pd.DataFrame(columns=['nc_month','m_novelty','m_usefulness','m_Noveltys2p1','m_Noveltys2p2','m_Noveltys2p3','m_Usefulnesss2p1','m_Usefulnesss2p2','m_Usefulnesss2p3','mavg_added','mavg_deleted','mavg_parents','mavg_files','c_count','c_contributors','documentation','peer','bug','feature','test'])

    return df2

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 



    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df = df.applymap(lambda x: x.encode('unicode_escape').
                 decode('utf-8') if isinstance(x, str) else x)
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save() 
    
    
def main():
    pd.options.display.max_rows = 10
    pd.options.display.float_format = '{:.3f}'.format

    df_commiters= pd.read_excel(COMMITERS_XLSX, sep=",",error_bad_lines=False,header=0,  encoding = "Latin1")

    for COMMIT_XLSX in REPOCOMMIT_LIST:
        print(COMMIT_XLSX)
        df_commit= pd.read_excel(COMMIT_XLSX, sep=",",error_bad_lines=False,header=0,  encoding = "Latin1")
        df_commit.drop(df_commit[(df_commit.PINDEX.notna()) & (df_commit.PINDEX.shift(-1).notna())].index, inplace= True)
        #clean_df(df_commit)
        df_commit.drop(columns=["Unnamed: 0"], inplace = True)
        
        #get month information for each commit
        df_commit = df_commit.assign( c_month =   pd.DatetimeIndex(pd.to_datetime(df_commit['OWNER'], infer_datetime_format=True,errors='coerce')).month)
        df_commit = df_commit.assign( c_day =   pd.DatetimeIndex(pd.to_datetime(df_commit['OWNER'], infer_datetime_format=True,errors='coerce')).day)
        #consolidate novelty and usefulness
        df_commit['con_novelty'] = df_commit.apply(consolidate_prob, args =('Noveltys2p1','Noveltys2p2','Noveltys2p3'), axis = 1)
        df_commit['con_usefulness'] = df_commit.apply(consolidate_prob, args =('Usefulnesss2p1','Usefulnesss2p2','Usefulnesss2p3'), axis = 1)
        
        #aggregate to monthly
        
        #get commits of a repo
        repo_commits = pd.DataFrame()
        write_commits = pd.DataFrame()
        indx = df_commit.columns
        indx = indx.append(pd.Index(['end_colab','end_cont','event_month','month','start_colab','start_cont','m_novelty','m_usefulness','m_Noveltys2p1','m_Noveltys2p2','m_Noveltys2p3','m_Usefulnesss2p1','m_Usefulnesss2p2','m_Usefulnesss2p3','mavg_added','mavg_deleted','mavg_parents','mavg_files','c_count','c_contributors','documentation','peer','bug','feature','test']))
        
    
        repo_df = df_commit.dropna(subset=['PINDEX'])
    
        # df_commit= df_commit.assign(n_REPO_ID = lambda x : "" if pd.isna(x['PINDEX']) else x['REPO_ID'])
        df_commit['n_REPO_ID'] = np.where(pd.isna(df_commit['PINDEX']), np.nan, df_commit['REPO_ID'])
        
        df_commit['n_REPO_ID'] = df_commit['n_REPO_ID'].fillna(method='ffill')
        
        df_commiters[0] = df_commiters[0].fillna(method='ffill')
        
        for i,row in repo_df.iterrows():
            print(row['REPO_ID'])
            repo_commits = df_commit[df_commit['n_REPO_ID']==row['REPO_ID']]
            
            repo_commits = repo_commits.iloc[1:]
    
            df_mcommit = consolidate_commits(repo_commits)
            write_commits = write_commits.append(row, sort = False, ignore_index = True)

            
            # get the monthwise commiters for the repo 
            repo_commiters = df_commiters[df_commiters[0].astype(int).astype(str)==row['REPO_ID']]
            
            repo_commiters = repo_commiters[['end_colab','end_cont','event_month','month','start_colab','start_cont']]
            
            repo_commiters = repo_commiters.iloc[1:]  
            if df_mcommit.shape[0] > 0:
                repo_commiters = pd.concat([repo_commiters.set_index('month'),df_mcommit.set_index('nc_month')], axis=1).reset_index().rename(columns={'index':'month'}) 
            else:
                print("CHECK REPO FOR COMMITS ", row['REPO_ID'])
               
                  
            write_commits = write_commits.append(repo_commiters, sort = False, ignore_index = True)                               

        write_commits = write_commits.reindex(indx, axis=1)
        write_commits = write_commits.drop(axis=1,columns=['REPO_ID.1', 'yhat','opt_deg_sup_ind','opt_deg_sup_org','Unnamed: 104','UNKNOWN','c_month','c_day','con_novelty','con_usefulness','Noveltys2p1',	'Noveltys2p2',	'Noveltys2p3',	'Usefulnesss2p1',	'Usefulnesss2p2',	
                                                           'Usefulnesss2p3','PUSHED_0917', 'STARS_0917', 'SUBSCRIBERS_0917', 'FORKS_0917', 'SIZE_0917', 'LICENCE_0917','STARS_1017', 'SUBSCRIBERS_1017', 'FORKS_1017', 'SIZE_1017', 'LICENCE_1017','int_sup', 
                                                           'int_sup_sq', 'month_jan_flag', 'month_feb_flag', 'month_mar_flag', 'month_apr_flag', 'month_may_flag', 'month_jun_flag','script_size_kb', 'script_size_mb', 'logsize_kb'])
        append_df_to_excel(MC_XLSX, write_commits,index=False)
        print("number of rows : ", write_commits.shape[0])
        
        # Save the full labeled data sample post processing in CSV
   
if __name__ == '__main__':
  main()
  