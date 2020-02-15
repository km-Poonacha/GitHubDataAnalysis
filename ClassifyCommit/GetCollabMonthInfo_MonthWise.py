# -*- coding: utf-8 -*-
"""
Created on Fri Jan 31 15:11:32 2020

@author: pmedappa

New code to find the collaborators for each project and date of their first contribution.This code also groups collaborators
,contributors and contributions monthwise.
Using the merge events sheet C:/Users\pmedappa\Dropbox\HEC\Project 5 - Roles and Coordination\Data\MergeEvents 
"""

import openpyxl
import pandas as pd
import numpy as np
import ast


EVENT_CSV ="C:/Users/pmedappa/Dropbox/HEC/Project 5 - Roles and Coordination/Data/MergeEvents/NewEvent2014_15_"

COL_MC_XLSX = "C:/Data/092019 CommitInfo/Contributors_monthwise/COL_MC_RepoCommit.xlsx"
NEW_XLSX = "C:/Data/092019 CommitInfo/Contributors_monthwise/Final_COL_MC_RepoCommit.xlsx"
DT_ERROR_LOG = "C:/Data/092019 CommitInfo/Contributors_monthwise/DT_ERROR_LOG.xlsx"

def monthwise(contri):
    #get month information for each commit
    contri.columns = ['NAN', 'C_TYPE','C_NAME','CONTRIBUTIONS','PULLS','S_DATE','E_DATE']

    contri= contri.assign( s_month =   pd.to_numeric(contri['S_DATE'].str.split('-').str[1]))
    # contri= contri.assign( s_month =   pd.DatetimeIndex(pd.to_datetime(contri['S_DATE'], infer_datetime_format=True,errors='coerce')).month)
    # contri = contri.assign( s_day =   pd.DatetimeIndex(pd.to_datetime(contri['S_DATE'], infer_datetime_format=True,errors='coerce')).day)
    
    contri= contri.assign( e_month =   pd.to_numeric(contri['E_DATE'].str.split('-').str[1]))
    # contri= contri.assign( e_month =   pd.DatetimeIndex(pd.to_datetime(contri['E_DATE'], infer_datetime_format=True,errors='coerce')).month)
    # contri = contri.assign( e_day =   pd.DatetimeIndex(pd.to_datetime(contri['E_DATE'], infer_datetime_format=True,errors='coerce')).day)

    contri['Collaborator']= contri['C_TYPE'].apply( lambda x: 1 if x == 'Collaborator' else 0)
    contri['Contributor']= contri['C_TYPE'].apply( lambda x: 1 if x == 'Contributor' else 0)
    df2 = pd.DataFrame({'month' : [1,2,3,4,5,6,7,8,9,10,11,12]})
    df2 = df2.set_index('month')
    df2 = pd.concat([df2,contri.groupby('s_month')['Collaborator'].sum().to_frame('start_colab')], axis=1)
    df2 = pd.concat([df2, contri.groupby('e_month')['Collaborator'].sum().to_frame('end_colab')], axis=1)
    df2 = pd.concat([df2,contri.groupby('s_month')['Contributor'].sum().to_frame('start_cont')], axis=1)
    df2 = pd.concat([df2, contri.groupby('e_month')['Contributor'].sum().to_frame('end_cont')], axis=1)
    df2 = df2.fillna(0)

    # df2.rename(columns={"index": "month"}, inplace = True)

    return df2

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 



    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
    
def main():
    """Get data from the training sample CSV and perform various cleaning and data preprocessing"""

    pd.options.display.max_rows = 10
    pd.options.display.float_format = '{:.3f}'.format

    for iterate in range(1,8):
        print("******** File ", iterate," ********")

        e_df = pd.read_csv(EVENT_CSV+str(iterate)+".csv", sep=",",error_bad_lines=False,header= None, low_memory=False, encoding = "Latin1")
        repo_df = e_df[e_df[0].notna()]
        e_df[0] = e_df[0].fillna(method='ffill')
        write_xl = pd.DataFrame()
        month_colab = pd.DataFrame()
        e_log = pd.DataFrame()
        print("NUmber of repos : ", repo_df.shape[0])

        for i, repo in repo_df.iterrows():
            # print(repo)
            write_xl = write_xl.append(repo, sort = False, ignore_index = True)
            month_colab = month_colab.append(repo, sort = False, ignore_index = True)
            events = e_df[e_df[0]==repo[0]]
            events= events.drop(events[(events[1] != '2014')].index)

            temp_df = pd.DataFrame()

            # events[4] = pd.to_datetime(events[4], format='%Y-%m-%d %H:%M:%S')
            events['event_month'] = pd.to_numeric(events[4].str.split('-').str[1])
            e_log = e_log.append(events[events['event_month'].isna()],sort = False, ignore_index = True)
            
            # try:
            #     events[4] = pd.to_datetime(events[4], format='%Y-%m-%d %H:%M:%S')
            # except: 
            #     try: events[4] = pd.to_datetime(events[4], format='%d/%m/%Y %H:%M:%S')
            #     except: 
            #         try: events[4] = pd.to_datetime(events[4], format='%d/%m/%y %H:%M:%S')
            #         except: pass

                
            # events[5] = pd.to_datetime(events[5], format='%Y-%m-%d %H:%M:%S')
            contributors = events[2].unique()
    
            write_events = events.drop(events[events[3] == 'PullRequestEvent'].index)
    
            collaborators = write_events[2].unique()
            # print("Repo : ",repo[0]," contributors ",len(contributors)," collaborators ",len(collaborators))
            
            # find first occurance of collborators
            for c in collaborators:
                c_events = write_events.drop(write_events[write_events[2] != c].index)
                fl_events = c_events[4].iloc[[0, -1]]
                write_xl = write_xl.append(pd.Series(["","Collaborator", c ,events.drop(events[events[2] != c].index).shape[0],c_events.shape[0], c_events[4].iloc[[0]].values[0], c_events[4].iloc[[-1]].values[0] ]), sort = False, ignore_index = True)
                temp_df = temp_df.append(pd.Series(["","Collaborator", c ,events.drop(events[events[2] != c].index).shape[0],c_events.shape[0], c_events[4].iloc[[0]].values[0], c_events[4].iloc[[-1]].values[0] ]), sort = False, ignore_index = True)
                contributors = contributors[contributors != c]
    
            for co in contributors:
                c_events = events.drop(events[events[2] != co].index)
                fl_events = c_events[4].iloc[[0, -1]]
                write_xl = write_xl.append(pd.Series(["","Contributor",  co , c_events.shape[0],0, c_events[4].iloc[[0]].values[0], c_events[4].iloc[[-1]].values[0] ]), sort = False, ignore_index = True)
                temp_df = temp_df.append(pd.Series(["","Contributor", c ,events.drop(events[events[2] != c].index).shape[0],c_events.shape[0], c_events[4].iloc[[0]].values[0], c_events[4].iloc[[-1]].values[0] ]), sort = False, ignore_index = True)
            
            
            # events['e_month'] = pd.DatetimeIndex(events[4]).month  
            
            month_contri = pd.DataFrame({'month' : [1,2,3,4,5,6,7,8,9,10,11,12]})
            month_contri = month_contri.set_index('month')
            month_contri = pd.concat([month_contri, events.groupby('event_month')['event_month'].count()], axis=1)
            month_contri = month_contri.fillna(0)
            month_data = pd.concat([monthwise(temp_df),month_contri],axis=1)
            month_data = month_data.reset_index()
            month_data.rename(columns={"index": "month"}, inplace = True)

            month_colab = pd.concat([month_colab,month_data],sort = False)
        
  
        append_df_to_excel(COL_MC_XLSX, write_xl,index=False)  
        append_df_to_excel(NEW_XLSX, month_colab, index=False)
        append_df_to_excel(DT_ERROR_LOG, e_log, index=False)  

if __name__ == '__main__':
  main()
  